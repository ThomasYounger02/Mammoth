# 表格读写

本节课会学习四个项目案例代码，实现对【10月员工绩效表】工作簿的各类“花式”读写操作。

不过呢万变不离其宗，完成一个项目问题的代码，首先需要我们**将“无限”的读写问题分解，对应到“有限”的基本模式**。

1.学习分解不同的Excel文件读写问题，熟悉四种

**Excel文件读写的问题场景模式**

①单元格读写、②按行读写、③按行取数计算、④按行取数存为字典

### 案例一：读取单元格的数据，原样写入，其他已有的工作簿，属于“单元格读写”模式。

```python
# 从openpyxl库导入load_workbook函数
from openpyxl import load_workbook

# 打开【10月员工绩效表】的工作簿，获取活动工作表
performance_wb  = load_workbook('./material/10月员工绩效表.xlsx')
performance_ws = performance_wb.active

# 打开【江宇工资信息表】的工作簿，获取活动工作表
info_wb = load_workbook('./material/江宇工资信息表.xlsx')
info_ws = info_wb.active

# 获取【绩效】值
performance = performance_ws['D14'].value
# 获取【奖金】值
bonus = performance_ws['E14'].value
# 获取【基本工资】值
base = performance_ws['F14'].value

# 写入【绩效】值
info_ws['E11'].value = performance
# 写入【奖金】值
info_ws['F11'].value = bonus
# 写入【基本工资】值
info_ws['G11'].value = base

# 保存对【江宇工资信息表】工作簿的写入
info_wb.save('./material/江宇工资信息表.xlsx')
```

!https://s3-us-west-2.amazonaws.com/secure.notion-static.com/6115a1c6-eb50-4781-8653-f3106a463fd9/Untitled.png

### 案例二：生成前十行绩效信息表.获取的数据是矩形区域的，一般要按行取出；然后将所取数据原样写入到另外的表格中。属于典型的按行读写模式。

```python
# 从openpyxl库导入load_workbook和Workbook
from openpyxl import load_workbook, Workbook

# 打开【10月员工绩效表.xlsx】工作簿
performance_wb = load_workbook('./material/10月员工绩效表.xlsx')
# 获取活动工作表
performance_ws = performance_wb.active

# 新建工作簿
new_wb = Workbook()
# 获取活动工作表
new_ws = new_wb.active

# 获取performance_ws的前十行数据
for row in performance_ws.iter_rows(max_row=10, values_only=True):
    # 将数据写入新的工作表
    new_ws.append(row)

# 保存新工作簿为【员工绩效表-模板.xlsx】
new_wb.save('员工绩效表-模板.xlsx')
```

!https://s3-us-west-2.amazonaws.com/secure.notion-static.com/0624b9c1-599b-4ffb-933a-ba64f30e93f0/Untitled.png

### 1. 获取哪些数据范围

根据要获取**什么数据范围内的单元格**来划分，大致可以分为三类：

1、已知坐标的个别单元格

2、单行或单列范围内的单元格

3、多行多列组成的矩形范围内的单元格

!https://s3-us-west-2.amazonaws.com/secure.notion-static.com/c2e6fa16-a880-43c5-81f6-b9ecca000aeb/Untitled.png

!https://s3-us-west-2.amazonaws.com/secure.notion-static.com/5343018f-a947-4e2e-8fc0-585dadd40c9c/Untitled.png

### 2. 如何使用数据

那么除了原样的写入数据，我们还可以从行数据中抽取数据，进行**计算**（数学计算、字符串拼接等），比如通过“提成”与“绩效”之和，得到本月所有奖金的金额。

另外，还可以抽取数据，**形成新的数据行**，比如抽取出每行第1个和最后1个单元格的内容，得到一行新的数据。

当然还可以将获得的数据，**存储为Python中的数据类型**，以便后续使用。比如把每行数据存为字典类型等等。

!https://s3-us-west-2.amazonaws.com/secure.notion-static.com/e3aac8f1-01d9-41f7-924c-008dddc6db3b/Untitled.png

### 3. 如何输出结果

主要是写入到其他工作表，输出新的工作簿内容

!https://s3-us-west-2.amazonaws.com/secure.notion-static.com/50292bcb-d63b-44b9-9a90-273b76bef7a7/Untitled.png

还可以“指定”从哪一行写起，“指定”写到具体坐标位置等；还可以写到不同的工作表对象中，比如原工作表、其他已有工作表、或者新建工作簿的工作表等。

另外，如果加上**“循环”**，还可以同时写到**多个不同**的工作表中，或者多个不同的坐标位置上。

!https://s3-us-west-2.amazonaws.com/secure.notion-static.com/4b9dee88-b2f9-46c6-9c84-cf8124e8cbb3/Untitled.png

!https://s3-us-west-2.amazonaws.com/secure.notion-static.com/607a2477-263f-483f-8cb5-3f3aaff87519/Untitled.png

### 案例三：计算并打印奖金信息。按行取数计算

```python
# 从openpyxl库导入load_workbook和Workbook
from openpyxl import load_workbook, Workbook

# 打开【10月员工绩效表.xlsx】工作簿
performance_wb = load_workbook('./material/10月员工绩效表.xlsx')
# 获取活动工作表
performance_ws = performance_wb.active

# 获取performance_ws中除表头外的数据
for row in performance_ws.iter_rows(min_row=2, values_only=True):
    # 读取【工号】
    staff_id = row[0]
    # 读取【员工姓名】
    staff_name = row[1]
    # 读取【绩效】
    performance = row[3]
    # 读取【提成】
    bonus = row[4]
    # 计算“奖金”
    award = performance + bonus
    # 打印结果
    print('工号：{}，姓名：{}，本月奖金为：{}'.format(staff_id, staff_name, award))
```

对values_only=True的说明。

上图中左侧的代码中，参数值为True，表示“只取单元格的值”，后续使用索引取到的，直接就是单元格的值。后续代码中不需**`工作表.value`**语句。

右侧的代码中，默认为False，表示不是“只取单元格的值”，那么使用索引取出的不仅仅是单元格值，而是单元格对象，因此后续的代码都需要增加**`.value`**获取单元格的值。

### 案例四：创建薪资信息字典。按行读取为字典

```python
# 从openpyxl库导入load_workbook函数
from openpyxl import load_workbook

# 打开【10月员工绩效表.xlsx】工作簿
performance_wb = load_workbook('./material/10月员工绩效表.xlsx')
# 获取活动工作表
performance_ws = performance_wb.active

# 创建员工信息字典
staff_info = {}

# 从第二行开始读取工作表中的信息
for row in performance_ws.iter_rows(min_row=2, values_only=True):
    # 取出工号
    member_number = row[0]
    # 将信息存入员工信息字典
    staff_info[member_number] = {
         '姓名': row[1],
         '部门': row[2],
         '绩效': row[3],
         '奖金': row[4],
         '基本工资': row[5],
         '是否确认': row[6]
     }
print(staff_info)
```

## 总结